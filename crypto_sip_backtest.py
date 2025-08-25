import ccxt
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from tqdm import tqdm

CRYPTO_LIST = {
    'BTC': 'BTC/USDT',
    'ETH': 'ETH/USDT',
    'XRP': 'XRP/USDT',
    'SOL': 'SOL/USDT',
    'BNB': 'BNB/USDT',
    'LTC': 'LTC/USDT',
    'DOGE': 'DOGE/USDT',
    'LINK': 'LINK/USDT',
    'ADA': 'ADA/USDT',
    'SUI': 'SUI/USDT'
}

SIP_AMOUNT = 20 / len(CRYPTO_LIST)  # USD amount to invest daily
BASE_EXCEL_FILE = 'crypto_sip_report.xlsx'
NUM_DAYS = 1000  # Number of days to backtest

exchange = ccxt.binance()

def SIP_backtest(CRYPTO, SIP_AMOUNT, timeframe, limit):
    data = exchange.fetch_ohlcv(CRYPTO, timeframe, limit=limit)
    df = pd.DataFrame(data, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
    df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms').dt.date  # Keep only date part
    df.set_index('timestamp', inplace=True)

    # create a new column with avg price for the day
    df['average'] = df[['open', 'high', 'low', 'close']].mean(axis=1)
    # calculate the percentage change in a new column
    df['percentage_change'] = df['close'].pct_change()

    # Calculate how much crypto is bought each day with SIP_AMOUNT at the close price
    df['crypto_bought'] = SIP_AMOUNT / df['close']

    # Calculate cumulative crypto holdings
    df['cumulative_crypto'] = df['crypto_bought'].cumsum()

    # Track cumulative investment: SIP_AMOUNT per day, so 15, 30, 45, ...
    df['cumulative_investment'] = SIP_AMOUNT * (df.reset_index().index + 1).values

    # Calculate cumulative portfolio value in USD at each day's close price
    df['portfolio_value'] = df['cumulative_crypto'] * df['close']

    # Calculate portfolio percentage change based on cumulative investment and portfolio value
    df['portfolio_pct_change'] = ((df['portfolio_value'] - df['cumulative_investment']) / df['cumulative_investment']) * 100
    df['portfolio_pct_change'] = df['portfolio_pct_change'].apply(lambda x: f"+{x:.2f}%" if x > 0 else f"{x:.2f}%")

    # Add a numeric column for sorting/filtering
    df['portfolio_pct_change_value'] = ((df['portfolio_value'] - df['cumulative_investment']) / df['cumulative_investment']) * 100

    # Round all decimal columns to 2 decimals
    decimal_cols = ['open', 'high', 'low', 'close', 'volume', 'average', 'percentage_change',
                    'crypto_bought', 'cumulative_crypto', 'cumulative_investment',
                    'portfolio_value', 'portfolio_pct_change_value']
    df[decimal_cols] = df[decimal_cols].round(2)

    return df

FREQUENCIES = {
    "daily": ("1d", NUM_DAYS),
    "weekly": ("1w", int(NUM_DAYS/7)),
    "monthly": ("1m", int(NUM_DAYS/30))
}
# Write each crypto's DataFrame to a separate sheet in the same Excel file

for freq_str, (timeframe, limit) in FREQUENCIES.items():
    EXCEL_FILE = f'{BASE_EXCEL_FILE.split(".")[0]}_{freq_str}.xlsx'
    summary_data = []

    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        sheet_names = []
        for name, symbol in tqdm(CRYPTO_LIST.items()):
            df = SIP_backtest(symbol, SIP_AMOUNT, timeframe, limit)
            df.to_excel(writer, sheet_name=name)
            sheet_names.append(name)

            last_row = df.iloc[-1]
            summary_data.append({
                'Crypto': name,
                'Start Date': df.index[0],
                'End Date': df.index[-1],
                'Total Invested ($)': last_row['cumulative_investment'],
                'Portfolio Value ($)': last_row['portfolio_value'],
                'P/L ($)': round(last_row['portfolio_value'] - last_row['cumulative_investment'], 2),
                'P/L (%)': round(last_row['portfolio_pct_change_value'], 2)
            })

        # Dashboard summary
        summary_df = pd.DataFrame(summary_data)
        summary_df = summary_df[['Crypto', 'Start Date', 'End Date', 'Total Invested ($)', 'Portfolio Value ($)', 'P/L ($)', 'P/L (%)']]

        total_row = {
            'Crypto': 'TOTAL',
            'Start Date': '',
            'End Date': '',
            'Total Invested ($)': summary_df['Total Invested ($)'].sum(),
            'Portfolio Value ($)': summary_df['Portfolio Value ($)'].sum(),
            'P/L ($)': summary_df['P/L ($)'].sum(),
            'P/L (%)': round(
                (summary_df['Portfolio Value ($)'].sum() - summary_df['Total Invested ($)'].sum())
                / summary_df['Total Invested ($)'].sum() * 100, 2
            )
        }
        summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)

        # TOP-N rows
        crypto_only_df = summary_df[summary_df['Crypto'] != 'TOTAL'].copy()

        topn_rows = []
        for n in range(1, len(crypto_only_df) + 1):
            topn = crypto_only_df.head(n)
            row = {
                'Crypto': f'TOP {n}',
                'Start Date': '',
                'End Date': '',
                'Total Invested ($)': topn['Total Invested ($)'].sum(),
                'Portfolio Value ($)': topn['Portfolio Value ($)'].sum(),
                'P/L ($)': topn['P/L ($)'].sum(),
                'P/L (%)': round(
                    (topn['Portfolio Value ($)'].sum() - topn['Total Invested ($)'].sum())
                    / topn['Total Invested ($)'].sum() * 100, 2
                )
            }
            topn_rows.append(row)

        empty_row = pd.DataFrame([{col: "" for col in summary_df.columns}])
        summary_df = pd.concat([summary_df, empty_row, pd.DataFrame(topn_rows)], ignore_index=True)
        summary_df.to_excel(writer, sheet_name='Dashboard', index=False)

        # Apply bold font to TOTAL row
        ws_dash = writer.book['Dashboard']
        total_row_idx = ws_dash.max_row - len(topn_rows)
        # total_row_idx_ = len(CRYPTO_LIST)+1
        for cell in ws_dash[total_row_idx-1]:
            cell.font = Font(bold=True)

        # Conditional formatting (red-yellow-green)
        workbook = writer.book
        for sheet_name in sheet_names:
            ws = workbook[sheet_name]
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'portfolio_pct_change_value':
                    col_letter = get_column_letter(idx)
                    ws.conditional_formatting.add(
                        f"{col_letter}2:{col_letter}{ws.max_row}",
                        ColorScaleRule(
                            start_type='num', start_value=-100, start_color='F8696B',  # Red
                            mid_type='num', mid_value=0, mid_color='FFEB84',          # Yellow
                            end_type='num', end_value=100, end_color='63BE7B'         # Green
                        )
                    )
                    break

        ws_dash = workbook['Dashboard']
        for idx, cell in enumerate(ws_dash[1], 1):
            if cell.value == 'P/L (%)':
                col_letter = get_column_letter(idx)
                ws_dash.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws_dash.max_row}",
                    ColorScaleRule(
                        start_type='num', start_value=-100, start_color='F8696B',
                        mid_type='num', mid_value=0, mid_color='FFEB84',
                        end_type='num', end_value=100, end_color='63BE7B'
                    )
                )
                break

        # Find P/L (%) column index
        pl_col_idx = None
        for idx, cell in enumerate(ws_dash[1], 1):
            if cell.value == "P/L (%)":
                pl_col_idx = idx
                break

        if pl_col_idx:
            crypto_rows = []
            topn_rows = []

            # Separate crypto rows and TopN rows
            for row in range(2, ws_dash.max_row + 1):
                label = ws_dash.cell(row=row, column=1).value
                val = ws_dash.cell(row=row, column=pl_col_idx).value
                if isinstance(val, (int, float)):
                    if label.startswith("TOP"):
                        topn_rows.append((label, val))
                    else:
                        crypto_rows.append((label, val))

            # --- Row 1: Best individual crypto ---
            if crypto_rows:
                best_crypto_label, best_crypto_val = max(crypto_rows, key=lambda x: x[1])
                row_idx = ws_dash.max_row + 2
                text = f"Best Crypto Performer: {best_crypto_label} with {best_crypto_val:.2f}% P/L"
                ws_dash.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=pl_col_idx)
                cell = ws_dash.cell(row=row_idx, column=1)
                cell.value = text
                cell.fill = PatternFill(start_color="FFF176", end_color="FFF176", fill_type="solid")
                cell.font = Font(bold=True)

            # --- Row 2: Best among TopN ---
            if topn_rows:
                best_topn_label, best_topn_val = max(topn_rows, key=lambda x: x[1])
                row_idx = ws_dash.max_row + 2
                text = f"Best Basket Performer: {best_topn_label} with {best_topn_val:.2f}% P/L"
                ws_dash.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=pl_col_idx)
                cell = ws_dash.cell(row=row_idx, column=1)
                cell.value = text
                cell.fill = PatternFill(start_color="FFF176", end_color="FFF176", fill_type="solid")
                cell.font = Font(bold=True)

        # Auto-fit columns
        for ws in workbook.worksheets:
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                col_letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[col_letter].width = length + 2

        workbook.save(EXCEL_FILE)
        print(f"Saved {EXCEL_FILE}")
