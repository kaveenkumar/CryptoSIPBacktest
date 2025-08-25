import warnings
warnings.filterwarnings("ignore")

import yfinance as yf
import pandas as pd
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule
import os

# -----------------------------
# Config
# -----------------------------
STOCK_LIST = {
    'AAPL': 'Apple',
    'MSFT': 'Microsoft',
    'AMZN': 'Amazon',
    'GOOGL': 'Alphabet',
    'META': 'Meta',
    'NVDA': 'NVIDIA',
    'TSLA': 'Tesla'
}

TOTAL_SIP_PER_PERIOD = 20.0  # split equally across stocks, per buy
BASE_EXCEL_FILE = "stocks_sip_report.xlsx"

# yfinance period/interval per frequency
FREQUENCIES = {
    "daily":  ("1d",  "3y"),
    "weekly": ("1wk", "3y"),
    "monthly":("1mo", "3y")
}

# -----------------------------
# Helpers
# -----------------------------
def fetch_ohlc(ticker: str, interval: str, period: str) -> pd.DataFrame:
    """
    Fetch OHLCV from Yahoo Finance; return DataFrame with Date as index (date only) and flat columns.
    """
    df = yf.download(ticker, interval=interval, period=period, auto_adjust=False, progress=False)
    if df is None or df.empty:
        return pd.DataFrame()
    # Ensure flat columns and date index (date only)
    df = df.reset_index()
    df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]
    # Standardize column names just in case
    rename_map = {"Adj Close": "AdjClose"}
    df = df.rename(columns=rename_map)
    # Keep date only (no time) for index
    df["Date"] = pd.to_datetime(df["Date"]).dt.date
    df = df.set_index("Date")
    # Ensure required columns exist
    for needed in ["Open", "High", "Low", "Close", "Volume"]:
        if needed not in df.columns:
            df[needed] = pd.NA
    return df[["Open", "High", "Low", "Close", "Volume"]].dropna()

def sip_backtest(df: pd.DataFrame, sip_amount: float) -> pd.DataFrame:
    """
    Simulate SIP on each bar's Close price.
    Mirrors your crypto logic and columns.
    """
    if df.empty:
        return pd.DataFrame()

    out = df.copy()
    # Average price column (for info)
    out["average"] = out[["Open", "High", "Low", "Close"]].mean(axis=1)
    # Daily/weekly/monthly percentage change on close
    out["percentage_change"] = out["Close"].pct_change()

    # Shares bought each period at Close
    out["shares_bought"] = sip_amount / out["Close"]

    # Cumulative shares and investment
    out["cumulative_shares"] = out["shares_bought"].cumsum()
    out["cumulative_investment"] = sip_amount * (pd.Series(range(1, len(out) + 1), index=out.index).values)

    # Portfolio value at Close
    out["portfolio_value"] = out["cumulative_shares"] * out["Close"]

    # Portfolio percentage change
    pct_val = ((out["portfolio_value"] - out["cumulative_investment"]) / out["cumulative_investment"]) * 100
    out["portfolio_pct_change_value"] = pct_val
    out["portfolio_pct_change"] = pct_val.apply(lambda x: f"+{x:.2f}%" if x > 0 else f"{x:.2f}%")

    # Round numeric columns
    decimal_cols = ["Open", "High", "Low", "Close", "Volume", "average", "percentage_change",
                    "shares_bought", "cumulative_shares", "cumulative_investment",
                    "portfolio_value", "portfolio_pct_change_value"]
    for col in decimal_cols:
        if col in out.columns:
            out[col] = out[col].astype(float).round(2)

    return out

def add_color_scale_percent(ws, header_label: str):
    """
    Apply red-yellow-green 3-color scale on a column with header == header_label.
    Red for negative, Yellow at 0, Green for positive. Uses fixed -100..0..+100 scale.
    """
    # locate header column index
    col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == header_label:
            col_idx = idx
            break
    if not col_idx:
        return
    col_letter = get_column_letter(col_idx)
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{ws.max_row}",
        ColorScaleRule(
            start_type='num', start_value=-100, start_color='F8696B',   # red
            mid_type='num',   mid_value=0,    mid_color='FFEB84',       # yellow
            end_type='num',   end_value=100,  end_color='63BE7B'        # green
        )
    )

def build_dashboard_rows(per_asset: list) -> pd.DataFrame:
    """
    per_asset is a list of dicts like:
      {'Stock': 'AAPL', 'Start Date': ..., 'End Date': ..., 'Total Invested ($)': ..., 'Portfolio Value ($)': ..., 'P/L ($)': ..., 'P/L (%)': ...}
    Returns summary_df with TOTAL, blank row, and TOP-N appended.
    """
    summary_df = pd.DataFrame(per_asset)
    summary_df = summary_df[['Stock', 'Start Date', 'End Date',
                             'Total Invested ($)', 'Portfolio Value ($)', 'P/L ($)', 'P/L (%)']]

    # TOTAL row
    total_row = {
        'Stock': 'TOTAL',
        'Start Date': '',
        'End Date': '',
        'Total Invested ($)': summary_df['Total Invested ($)'].sum(),
        'Portfolio Value ($)': summary_df['Portfolio Value ($)'].sum(),
        'P/L ($)': summary_df['P/L ($)'].sum(),
        'P/L (%)': round(
            (summary_df['Portfolio Value ($)'].sum() - summary_df['Total Invested ($)'].sum())
            / max(summary_df['Total Invested ($)'].sum(), 1e-12) * 100, 2
        )
    }
    summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)

    # TOP-N (sort by P/L (%) descending)
    crypto_only_df = summary_df[summary_df['Stock'] != 'TOTAL'].copy()
    crypto_only_df = crypto_only_df.sort_values(by='P/L (%)', ascending=False)

    topn_rows = []
    for n in range(1, len(crypto_only_df) + 1):
        topn = crypto_only_df.head(n)
        topn_rows.append({
            'Stock': f'TOP {n}',
            'Start Date': '',
            'End Date': '',
            'Total Invested ($)': topn['Total Invested ($)'].sum(),
            'Portfolio Value ($)': topn['Portfolio Value ($)'].sum(),
            'P/L ($)': topn['P/L ($)'].sum(),
            'P/L (%)': round(
                (topn['Portfolio Value ($)'].sum() - topn['Total Invested ($)'].sum())
                / max(topn['Total Invested ($)'].sum(), 1e-12) * 100, 2
            )
        })

    # blank row between TOTAL and TOP-N
    empty_row = pd.DataFrame([{col: "" for col in summary_df.columns}])
    summary_df = pd.concat([summary_df, empty_row, pd.DataFrame(topn_rows)], ignore_index=True)
    return summary_df

def add_summary_banners(ws):
    """
    Add two merged yellow banners:
      1) Best individual stock performer
      2) Best TopN performer
    """
    # locate 'P/L (%)' column
    pl_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'P/L (%)':
            pl_col_idx = idx
            break
    if not pl_col_idx:
        return

    max_col = ws.max_column

    # Separate crypto rows (stocks) vs TOP rows (skip TOTAL & blanks)
    crypto_rows = []
    topn_rows = []
    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        plval = ws.cell(row=r, column=pl_col_idx).value
        if isinstance(plval, (int, float)):
            if isinstance(label, str) and label.startswith("TOP"):
                topn_rows.append((label, plval))
            elif label not in ("TOTAL", "", None):
                crypto_rows.append((label, plval))

    # Banner 1: Best individual stock
    next_row = ws.max_row + 2
    if crypto_rows:
        best_stock, best_val = max(crypto_rows, key=lambda x: x[1])
        text = f"Best Stock Performer: {best_stock} with {best_val:.2f}% P/L"
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=max_col)
        c = ws.cell(row=next_row, column=1)
        c.value = text
        c.fill = PatternFill(start_color="FFF176", end_color="FFF176", fill_type="solid")
        c.font = Font(bold=True)

    # Banner 2: Best TopN portfolio
    next_row = ws.max_row + 2
    if topn_rows:
        best_topn, best_val = max(topn_rows, key=lambda x: x[1])
        text = f"Best TopN Portfolio: {best_topn} with {best_val:.2f}% P/L"
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=max_col)
        c = ws.cell(row=next_row, column=1)
        c.value = text
        c.fill = PatternFill(start_color="FFF176", end_color="FFF176", fill_type="solid")
        c.font = Font(bold=True)

def autofit_all_columns(workbook):
    for ws in workbook.worksheets:
        for column_cells in ws.columns:
            try:
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            except:
                length = 15
            col_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[col_letter].width = min(length + 2, 60)

# -----------------------------
# Main
# -----------------------------
os.makedirs("results_stocks", exist_ok=True)
for freq_str, (interval, period) in FREQUENCIES.items():
    excel_file = f"results_stocks/{BASE_EXCEL_FILE.split('.')[0]}_{freq_str}.xlsx"
    per_asset_rows = []
    sheet_names = []

    # SIP per stock = TOTAL_SIP_PER_PERIOD / N
    sip_amount_per_stock = TOTAL_SIP_PER_PERIOD / len(STOCK_LIST)

    # Build Excel
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        # Each stock → sheet
        for ticker, name in tqdm(STOCK_LIST.items(), desc=f"Processing {freq_str}"):
            ohlc = fetch_ohlc(ticker, interval=interval, period=period)
            if ohlc.empty:
                # still record a minimal row with zeros to keep dashboard consistent
                per_asset_rows.append({
                    'Stock': ticker,
                    'Start Date': '',
                    'End Date': '',
                    'Total Invested ($)': 0.0,
                    'Portfolio Value ($)': 0.0,
                    'P/L ($)': 0.0,
                    'P/L (%)': 0.0
                })
                continue

            df = sip_backtest(ohlc, sip_amount=sip_amount_per_stock)
            # Save per-stock sheet
            df.to_excel(writer, sheet_name=ticker)
            sheet_names.append(ticker)

            # Collect summary (last row)
            last = df.iloc[-1]
            per_asset_rows.append({
                'Stock': ticker,
                'Start Date': df.index[0],
                'End Date': df.index[-1],
                'Total Invested ($)': float(last['cumulative_investment']),
                'Portfolio Value ($)': float(last['portfolio_value']),
                'P/L ($)': round(float(last['portfolio_value'] - last['cumulative_investment']), 2),
                'P/L (%)': round(float(last['portfolio_pct_change_value']), 2)
            })

        # Dashboard (summary + TOTAL + blank + TOP-N)
        dashboard_df = build_dashboard_rows(per_asset_rows)
        dashboard_df.to_excel(writer, sheet_name="Dashboard", index=False)

        # Workbook-level formatting
        wb = writer.book

        # Bold TOTAL row (it's right before the blank row we added)
        ws_dash = wb["Dashboard"]
        # Find TOTAL row by scanning column A
        total_row_idx = None
        for r in range(2, ws_dash.max_row + 1):
            if ws_dash.cell(row=r, column=1).value == "TOTAL":
                total_row_idx = r
                break
        if total_row_idx:
            for cell in ws_dash[total_row_idx]:
                cell.font = Font(bold=True)

        # Conditional format on each stock sheet for 'portfolio_pct_change_value'
        for s in sheet_names:
            ws = wb[s]
            add_color_scale_percent(ws, header_label="portfolio_pct_change_value")

        # Conditional format on Dashboard 'P/L (%)'
        add_color_scale_percent(ws_dash, header_label="P/L (%)")

        # Add two merged yellow banners at bottom
        add_summary_banners(ws_dash)

        # Autofit
        autofit_all_columns(wb)

        wb.save(excel_file)

    print(f"✅ Saved {excel_file}")
