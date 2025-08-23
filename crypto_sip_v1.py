import ccxt
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


CRYPTO_LIST = {
    'BTC': 'BTC/USDT',
    'ETH': 'ETH/USDT',
    'XRP': 'XRP/USDT',
    'SOL': 'SOL/USDT',
    'BNB': 'BNB/USDT',
    'LTC': 'LTC/USDT',
    'DOGE': 'DOGE/USDT',
    'LINK': 'LINK/USDT',
    'ADA': 'ADA/USDT',
    'SUI': 'SUI/USDT'
}
SIP_AMOUNT = 20 / len(CRYPTO_LIST)  # USD amount to invest daily
EXCEL_FILE = 'crypto_sip_report.xlsx'
NUM_DAYS = 1000  # Number of days to backtest

daily = True
weekly = not(daily)
monthly = False

exchange = ccxt.binance()

def SIP_backtest(CRYPTO, SIP_AMOUNT, daily, weekly, monthly):
    if daily:
        data = exchange.fetch_ohlcv(CRYPTO, '1d', limit=NUM_DAYS)
    elif weekly:
        data = exchange.fetch_ohlcv(CRYPTO, '1w', limit=int(NUM_DAYS/7))
    elif monthly:
        data = exchange.fetch_ohlcv(CRYPTO, '1m', limit=int(NUM_DAYS/30))

    df = pd.DataFrame(data, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
    df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms').dt.date  # Keep only date part
    df.set_index('timestamp', inplace=True)

    # create a new column with avg price for the day
    df['average'] = df[['open', 'high', 'low', 'close']].mean(axis=1)
    # calculate the percentage change in a new column
    df['percentage_change'] = df['close'].pct_change()

    # Calculate how much crypto is bought each day with SIP_AMOUNT at the close price
    df['crypto_bought'] = SIP_AMOUNT / df['close']

    # Calculate cumulative crypto holdings
    df['cumulative_crypto'] = df['crypto_bought'].cumsum()

    # Track cumulative investment: SIP_AMOUNT per day, so 15, 30, 45, ...
    df['cumulative_investment'] = SIP_AMOUNT * (df.reset_index().index + 1).values

    # Calculate cumulative portfolio value in USD at each day's close price
    df['portfolio_value'] = df['cumulative_crypto'] * df['close']

    # Calculate portfolio percentage change based on cumulative investment and portfolio value
    df['portfolio_pct_change'] = ((df['portfolio_value'] - df['cumulative_investment']) / df['cumulative_investment']) * 100
    df['portfolio_pct_change'] = df['portfolio_pct_change'].apply(lambda x: f"+{x:.2f}%" if x > 0 else f"{x:.2f}%")

    # Add a numeric column for sorting/filtering
    df['portfolio_pct_change_value'] = ((df['portfolio_value'] - df['cumulative_investment']) / df['cumulative_investment']) * 100

    # Round all decimal columns to 2 decimals
    decimal_cols = ['open', 'high', 'low', 'close', 'volume', 'average', 'percentage_change',
                    'crypto_bought', 'cumulative_crypto', 'cumulative_investment',
                    'portfolio_value', 'portfolio_pct_change_value']
    df[decimal_cols] = df[decimal_cols].round(2)

    return df

# Write each crypto's DataFrame to a separate sheet in the same Excel file
summary_data = []

# Determine frequency string for filename
if daily:
    freq_str = "daily"
elif weekly:
    freq_str = "weekly"
elif monthly:
    freq_str = "monthly"
EXCEL_FILE = f'{EXCEL_FILE.split(".")[0]}_{freq_str}.xlsx'

with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    sheet_names = []
    for name, symbol in CRYPTO_LIST.items():
        df = SIP_backtest(symbol, SIP_AMOUNT, daily, weekly, monthly)
        df.to_excel(writer, sheet_name=name)
        sheet_names.append(name)
        # Collect summary info: last date, total invested, final value, P/L, P/L %
        last_row = df.iloc[-1]
        summary_data.append({
            'Crypto': name,
            'Start Date': df.index[0],
            'End Date': df.index[-1],
            'Total Invested ($)': last_row['cumulative_investment'],
            'Portfolio Value ($)': last_row['portfolio_value'],
            'P/L ($)': round(last_row['portfolio_value'] - last_row['cumulative_investment'], 2),
            'P/L (%)': round(last_row['portfolio_pct_change_value'], 2)
        })

    # Create summary dashboard DataFrame
    summary_df = pd.DataFrame(summary_data)
    summary_df = summary_df[['Crypto', 'Start Date', 'End Date', 'Total Invested ($)', 'Portfolio Value ($)', 'P/L ($)', 'P/L (%)']]

    # Add a total row
    total_row = {
        'Crypto': 'TOTAL',
        'Start Date': '',
        'End Date': '',
        'Total Invested ($)': summary_df['Total Invested ($)'].sum(),
        'Portfolio Value ($)': summary_df['Portfolio Value ($)'].sum(),
        'P/L ($)': summary_df['P/L ($)'].sum(),
        'P/L (%)': round(
            (summary_df['Portfolio Value ($)'].sum() - summary_df['Total Invested ($)'].sum())
            / summary_df['Total Invested ($)'].sum() * 100, 2
        )
    }
    summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)
    summary_df.to_excel(writer, sheet_name='Dashboard', index=False)

    # Apply bold font to the TOTAL row in Dashboard
    ws_dash = writer.book['Dashboard']
    total_row_idx = ws_dash.max_row  # The last row is the TOTAL row
    for cell in ws_dash[total_row_idx]:
        cell.font = Font(bold=True)

    # Apply color scale formatting after saving all sheets
    workbook = writer.book

    # Color scale for each crypto sheet's portfolio_pct_change_value column
    for sheet_name in sheet_names:
        ws = workbook[sheet_name]
        # Find the column index for 'portfolio_pct_change_value'
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'portfolio_pct_change_value':
                col_letter = get_column_letter(idx)
                # Apply color scale (red for low, yellow for mid, green for high)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    ColorScaleRule(start_type='min', start_color='F8696B',
                                   mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                   end_type='max', end_color='63BE7B')
                )
                break

    # Color scale for Dashboard sheet's 'P/L (%)' column
    ws_dash = workbook['Dashboard']
    for idx, cell in enumerate(ws_dash[1], 1):
        if cell.value == 'P/L (%)':
            col_letter = get_column_letter(idx)
            ws_dash.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws_dash.max_row}",
                ColorScaleRule(start_type='min', start_color='F8696B',
                               mid_type='percentile', mid_value=50, mid_color='FFEB84',
                               end_type='max', end_color='63BE7B')
            )
            break

    # Auto-fit columns for all sheets
    for ws in workbook.worksheets:
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            col_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[col_letter].width = length + 2  # Add some padding

    # Save workbook with formatting
    workbook.save(EXCEL_FILE)
